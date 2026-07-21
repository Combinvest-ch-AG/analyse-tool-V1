"""Build compact browser data from the official BAG/Priminfo 2026 files.

Source files are intentionally kept out of the application bundle. Download them
to tmp/priminfo-2026 and run this script from the repository root.
"""

from __future__ import annotations

import csv
import json
from collections import defaultdict
from pathlib import Path

import openpyxl


ROOT = Path(__file__).resolve().parents[1]
SOURCE = ROOT / "tmp" / "priminfo-2026"
OUTPUT = ROOT / "data" / "priminfo-2026"


def clean(value: object) -> str:
    return " ".join(str(value or "").replace("\n", " ").split())


def insurer_names() -> dict[int, str]:
    workbook = openpyxl.load_workbook(
        SOURCE / "versicherer.xlsx", read_only=True, data_only=True
    )
    sheet = workbook["Index "]
    result: dict[int, str] = {}
    for row in sheet.iter_rows(min_row=4, values_only=True):
        number, name = row[1], row[3]
        if isinstance(number, (int, float)) and name:
            result[int(number)] = clean(name)
    return result


def locations() -> list[dict[str, object]]:
    workbook = openpyxl.load_workbook(
        SOURCE / "praemienregionen.xlsx", read_only=True, data_only=True
    )
    sheet = workbook["A_COM"]
    seen: set[tuple[object, ...]] = set()
    result: list[dict[str, object]] = []
    for row in sheet.iter_rows(min_row=6, values_only=True):
        bfs, canton, commune, region, _district, postal_code, locality = row[:7]
        if not all((bfs, canton, commune, region, postal_code, locality)):
            continue
        key = (int(bfs), str(canton), int(region), int(postal_code), clean(locality))
        if key in seen:
            continue
        seen.add(key)
        result.append(
            {
                "b": int(bfs),
                "c": str(canton),
                "g": clean(commune),
                "r": int(region),
                "p": int(postal_code),
                "o": clean(locality),
            }
        )
    return sorted(result, key=lambda item: (item["p"], item["o"], item["g"]))


def premium_files(names: dict[int, str]) -> dict[str, int]:
    grouped: dict[
        str,
        dict[tuple[str, str, int, str, str, str, str, str], dict[int, float]],
    ] = defaultdict(lambda: defaultdict(dict))
    path = SOURCE / "Praemien_CH.csv"
    with path.open(encoding="utf-8-sig", newline="") as handle:
        for row in csv.DictReader(handle):
            canton = row["Kanton"]
            if len(canton) != 2 or canton in {"ZE", "ZR"}:
                continue
            region = row["Region"].removeprefix("PR-REG CH")
            franchise = int(row["Franchise"].removeprefix("FRA-"))
            insurer = int(row["Versicherer"])
            key = (
                region,
                row["Altersklasse"].removeprefix("AKL-"),
                insurer,
                "MIT" if row["Unfalleinschluss"].startswith("MIT-") else "OHNE",
                row["Tariftyp"].removeprefix("TAR-"),
                row["Tarif"],
                clean(row["Tarifbezeichnung"]),
                row["Altersuntergruppe"],
            )
            grouped[canton][key][franchise] = float(row["Prämie"])

    counts: dict[str, int] = {}
    for canton, tariffs in grouped.items():
        compact = []
        for key, premiums in tariffs.items():
            region, age, insurer, accident, tariff_type, tariff, label, subgroup = key
            compact.append(
                {
                    "r": int(region),
                    "a": age,
                    "i": insurer,
                    "u": accident,
                    "y": tariff_type,
                    "t": tariff,
                    "n": label,
                    "s": subgroup,
                    "p": [[franchise, premium] for franchise, premium in sorted(premiums.items())],
                }
            )
        compact.sort(key=lambda x: (x["r"], x["a"], x["u"], names.get(x["i"], ""), x["n"], x["t"]))
        (OUTPUT / "premiums").mkdir(parents=True, exist_ok=True)
        (OUTPUT / "premiums" / f"{canton}.json").write_text(
            json.dumps(compact, ensure_ascii=False, separators=(",", ":")),
            encoding="utf-8",
        )
        counts[canton] = len(compact)
    return counts


def main() -> None:
    OUTPUT.mkdir(parents=True, exist_ok=True)
    names = insurer_names()
    locs = locations()
    counts = premium_files(names)
    (OUTPUT / "insurers.json").write_text(
        json.dumps(names, ensure_ascii=False, separators=(",", ":")), encoding="utf-8"
    )
    (OUTPUT / "locations.json").write_text(
        json.dumps(locs, ensure_ascii=False, separators=(",", ":")), encoding="utf-8"
    )
    meta = {
        "year": 2026,
        "premiumRows": sum(counts.values()),
        "locations": len(locs),
        "cantons": counts,
        "source": "Bundesamt für Gesundheit BAG / Priminfo",
        "premiumUrl": "https://opendata.swiss/de/dataset/health-insurance-premiums",
        "regionUrl": "https://www.priminfo.admin.ch/de/downloads/aktuell",
        "generatedBy": "scripts/build-priminfo-2026.py",
    }
    (OUTPUT / "meta.json").write_text(
        json.dumps(meta, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    print(json.dumps(meta, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
